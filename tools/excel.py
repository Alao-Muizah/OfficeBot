import openpyxl 
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import os 


# =========== Getting the absolute file path to use =========== #

def _resolve_path(path: str, workdir: str) -> str:
    if os.path.isabs(path):
        return path
    return os.path.join(workdir, path)


# ====================== Creating the Excel file ====================== #

def create_excel(path: str, sheet_name: str, workdir: str) -> dict:
    full_path = _resolve_path(path, workdir)

    parent_dir = os.path.dirname(full_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    wb.save(full_path)

    return {"status": "ok", 
            "message": f"Created '{path}' with sheet '{sheet_name}'", 
            "path": full_path}


# ====================== Reading the Excel Sheet ====================== #

def read_sheet(path:str, sheet_name:str, workdir:str) -> dict:

    full_path = _resolve_path(path, workdir)

    if not os.path.exists(full_path):
        return {"status": "error", "message": f"File not found: {full_path}"}
    
    wb = openpyxl.load_workbook(full_path)

    if sheet_name not in wb.sheetnames:
        return {"status": "error", 
                "message": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}
    
    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        return {"status": "ok", "data": [], "message": "Sheet is empty"}
    
    headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(all_rows[0])]
    
    data = []
    for row in all_rows[1:]:
        record = {headers[i]: row[i] for i in range(len(headers))}
        data.append(record)

    return {"status": "ok", "data": data,
             "headers": headers,
               "row_count": len(data)} 


# ====================== Writing headers to the excel file ====================== #

def write_headers(path: str, sheet_name:str, headers: list, workdir:str) -> dict:

    full_path = _resolve_path(path, workdir)

    if not os.path.exists(full_path):
        return {
            "status": "error",
            "message": f"File not found '{full_path}'"
        }
    
    wb = openpyxl.load_workbook(full_path)

    if sheet_name not in wb.sheetnames:
        return {
            "status": "error",
            "message": f"Sheet not found '{sheet_name}'"
        }
    
    ws = wb[sheet_name]

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D3D3D3")

    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    wb.save(full_path)
 
    return {
        "status": "ok",
        "message": f'Wrote {len(header)} headers to row 1'
    }


# ====================== writing rows to the Excel file ====================== #

def write_rows(path: str, sheet_name: str, rows: list, start_row: int, workdir: str) -> dict:
    full_path = _resolve_path(path, workdir)

    if not os.path.exists(full_path):
        return {
            "status": "error",
            "message": f"File not found '{sheet_name}'"
        }
    
    wb = openpyxl.load_workbook(full_path)

    if sheet_name not in wb.sheetnames :
        return {
            "status": "error",
            "message": f"Sheet '{sheet_name}' not found"
        }
    
    ws = wb[sheet_name]

    for row_idx, row_data in enumerate(rows):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row = start_row + row_idx, column=col_idx, value=value)

    wb.save(full_path)

    return {
        "status": "ok",
        "message": f"Wrote {len(rows)} rows starting from {start_row}",
    }

# ====================== Writing cells to the Excel file ====================== #

def write_cell(path: str, sheet_name:str, cell_ref: str, value:str, workdir:str) -> dict:

    full_path = _resolve_path(path, workdir)

    if not os.path.exists(full_path):
        return {
            "status": "error",
            "message": f"File not found '{full_path}"
        }
    
    wb = openpyxl.load_workbook(full_path)

    if sheet_name not in wb.sheetnames:
        return {
            "status": "error",
            "message": f"Sheet not found {sheet_name}"
        }
    
    ws = wb[sheet_name]
    ws[cell_ref] = value
   
    wb.save(full_path)

    return {
        "status": "ok",
        "message": f"Changed value at cell '{cell_ref}' to {value}"
    }


# ====================== Applying formula to the Excel file ====================== #

def apply_formula(path:str, sheet_name:str, cell_ref:str, formula:str, workdir:str) -> dict:

    
    full_path = _resolve_path(path, workdir)

    if not os.path.exists(full_path):
        return {
            "status": "error",
            "message": f"File not found '{full_path}'"
        }
    
    if not formula.startswith('='):
        return { 
            "status": "error",
            "message": "Formula must start with '='"
        }
    wb = openpyxl.load_workbook(full_path)

    if sheet_name not in wb.sheetnames:
        return {
            "status": "error",
            "message": f"Sheet '{sheet_name}' not found"
        }
    ws = wb[sheet_name]
    ws[cell_ref] = formula

    wb.save(full_path)

    return {
        "status": "ok",
        "mesage": f"Applied formula '{formula}' to '{cell_ref}'"
    }



# ====================== Formatting cell ====================== #

def format_cell(path:str, sheet_name:str, cell_ref:str, bold:bool, font_size: int, font_color:str, workdir:str) -> dict:

    full_path = _resolve_path(path, workdir)

    if not os.path.exists(full_path):
        return {
            "status": "error",
            "message": f"File not found '{full_path}'"
        }
    
    wb = openpyxl.load_workbook(full_path)
    
    if sheet_name not in wb.sheetnames:
        return {
            "status": "error",
            "message": f"Sheet '{sheet_name}' was not found"
        }
    ws = wb[sheet_name]
    cell = ws[cell_ref] 

    cell.font = Font(
        bold=bold,
        size=font_size,
        color=font_color.lstrip("#")
    )

    wb.save(full_path)

    return {
        "status": "ok",
        "message": f"Formatted {cell_ref}: bold={bold}, font-size = {font_size}, font color = {font_color}"
    }


# ====================== Setting column width ====================== #

def set_column_width(path:str, sheet_name:str, column:str, width:float, workdir:str) -> dict:

    full_path = _resolve_path(path, workdir)

    if not os.path.exists(full_path):
        return {
            "status": "error",
            "message": f"File not found '{full_path}'"
        }
    
    wb = openpyxl.load_workbook(full_path)
    
    if sheet_name not in wb.sheetnames:
        return {
            "status": "error",
            "message": f"Sheet '{sheet_name}' was not found"
        }
    ws = wb[sheet_name] 
    ws.column_dimensions[column.upper()].width = width 

    wb.save(full_path)

    return {
        "status": "ok",
        "message": f"Set column {column.upper()} width to {width}"
    }


# ====================== Adding new sheet ====================== #

def add_sheet(path:str, sheet_name:str, workdir:str) -> dict:

    full_path = _resolve_path(path, workdir)

    if not os.path.exists(full_path):
        return {
            "status": "error",
            "message": f"File not found '{full_path}'"
        }
    
    wb = openpyxl.load_workbook(full_path)
    
    if sheet_name in wb.sheetnames:
        return {
            "status": "error",
            "message": f"Sheet '{sheet_name}' already exists"
        }
    
    wb.create_sheet(title=sheet_name)
    wb.save(full_path)

    return {
        "status": "ok",
        "message": f"Added sheet '{sheet_name}'"
    }


# ====================== Listing sheets ====================== #

def list_sheets(path:str, workdir:str) -> dict:

    full_path = _resolve_path(path, workdir)

    if not os.path.exists(full_path):
        return {
            "status": "error",
            "message": f"File not found '{full_path}'"
        }
    
    wb = openpyxl.load_workbook(full_path)
    
    if not wb.sheetnames:
        return {
            "status": "error",
            "message": f"Workbook '{full_path}' is empty"
        }
    return {
        "status": "ok",
        "sheets": wb.sheetnames
    }

